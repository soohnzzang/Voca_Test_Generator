import os
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from os.path import splitext, basename

def create_excel_from_input():
    """사용자가 입력한 단어로 엑셀 파일을 생성"""
    word_list = []
    
    # 단어와 뜻 입력 받기
    while True:
        english_word = input("영어 단어를 입력하시오 (끝내려면 'end' 입력): ").strip()
        if english_word.lower() == 'end':
            break
        korean_meaning = input("한글 뜻을 입력하시오: ").strip()
        word_list.append((english_word, korean_meaning))
    
    # 엑셀 파일 생성
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "단어 리스트"
    
    # 서식 정의 (폰트, 배경색, 정렬 등)
    header_font = Font(name ="Pretendard Variable", bold=True, color="FFFFFF", size=14)  # 굵고 흰색 글씨
    header_fill = PatternFill(start_color="6091F2", end_color="6091F2", fill_type="solid")  # 검정색 배경
    alignment = Alignment(horizontal="center", vertical="center")  # 가운데 정렬
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # 얇은 테두리
    cell_font = Font(name="Pretendard Variable", size=13) 

    # 헤더 추가 및 서식 적용
    headers = ["영어 단어", "한글 뜻"]
    sheet.append(headers)
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # 단어 리스트 추가
    for row_idx, (word, meaning) in enumerate(word_list, start=2):
        sheet[f'A{row_idx}'] = word
        sheet[f'B{row_idx}'] = meaning

        # 각 셀에 서식 적용
        sheet[f'A{row_idx}'].alignment = alignment
        sheet[f'B{row_idx}'].alignment = alignment
        sheet[f'A{row_idx}'].border = thin_border
        sheet[f'B{row_idx}'].border = thin_border
    
    # 열 너비 자동 조정
    sheet.column_dimensions['A'].width = 20  # 영어 단어 열 너비
    sheet.column_dimensions['B'].width = 30  # 한글 뜻 열 너비

    # 저장할 파일 경로 입력받기
    output_file = input("엑셀 파일을 저장할 경로를 입력하세요 (예: C:/path/to/your_file.xlsx): ").strip('"')
    
    # 엑셀 파일 저장
    workbook.save(output_file)
    print(f"엑셀 파일이 {output_file}에 저장되었습니다.")
    
    return output_file  # 생성된 엑셀 파일 경로 반환

def create_voca_test():
    """엑셀 파일을 기반으로 테스트지를 생성"""
    # 파일 경로 입력
    file_path = input("새로운 단어 리스트 파일 경로를 입력하세요: ").strip('"')

    # 파일 경로에 백슬래시(\)가 있는지 확인하고 변환
    if "\\" in file_path:
        converted_path = file_path.replace("\\", "/")
        print(f"백슬래시를 슬래시로 변환했습니다: {converted_path}")
    else:
        converted_path = file_path
        print(f"경로에 변환이 필요 없습니다: {converted_path}")

    # 파일이 존재하는지 확인
    if not os.path.exists(converted_path):
        print(f"파일을 찾을 수 없습니다: {converted_path}")
        return
    
    # 엑셀 파일 분석
    workbook = load_workbook(converted_path)
    sheet = workbook.active
    
    print(f"파일이 성공적으로 로드되었습니다: {converted_path}")
    
    # 첫 번째 열은 영어 단어, 두 번째 열은 한글 뜻이라고 가정
    max_row = sheet.max_row
    
    # 1. 두 번째 열을 빈칸으로 처리 (한글 뜻을 빈칸으로)
    for row in range(2, max_row + 1):  # 2번째 행부터 처리
        sheet[f'B{row}'].value = ""  # 두 번째 열을 빈칸으로
    
    # 2. 첫 번째 열(영어 단어)에서 빈칸을 제외하고 셔플
    english_words = [sheet[f'A{row}'].value for row in range(2, max_row + 1) if sheet[f'A{row}'].value]  # 빈칸 제외
    random.shuffle(english_words)  # 영어 단어를 셔플
    
    # 셔플된 영어 단어를 다시 첫 번째 열에 입력
    for row in range(2, max_row + 1):
        if sheet[f'A{row}'].value:  # 영어 단어가 있을 때만 입력
            sheet[f'A{row}'].value = english_words.pop(0)  # 셔플된 단어를 입력
    
    # 새로운 폴더에 파일 생성
    output_dir = r"C:/soohnzzang/metaverse/토익-강의모음/voca_test/output"
    
    # 폴더가 없으면 생성
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 입력 파일명에서 확장자 제거 및 기본 이름 가져오기
    base_name, ext = splitext(basename(file_path))
    
    # 파일 이름 설정 및 넘버링 처리 (예: voca_test01-01.xlsx)
    file_number = 0
    output_file = os.path.join(output_dir, f"{base_name}-{file_number:02}{ext}")
    
    # 같은 이름의 파일이 존재할 경우 넘버링 증가
    while os.path.exists(output_file):
        file_number += 1
        output_file = os.path.join(output_dir, f"{base_name}-{file_number:02}{ext}")
    
    # 엑셀 파일 저장
    workbook.save(output_file)
    
    print(f"결과 파일이 {output_file}로 저장되었습니다.")

def main():
    """프로그램 시작 지점: 모드 선택"""
    mode = input("단어를 입력하여 엑셀을 생성하시겠습니까? (y/n): ").strip().lower()

    if mode == 'y':
        # 단어를 입력하여 엑셀 생성 모드
        excel_file_path = create_excel_from_input()
        print(f"엑셀 파일이 생성되었습니다: {excel_file_path}")
        create_voca_test()
    elif mode == 'n':
        # 기존 엑셀 파일 경로를 입력받아 테스트 생성 모드
        create_voca_test()
    else:
        print("잘못된 입력입니다. 프로그램을 종료합니다.")

# 프로그램 시작
if __name__ == "__main__":
    main()

